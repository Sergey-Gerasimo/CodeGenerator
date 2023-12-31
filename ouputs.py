
import math 
from openpyxl import Workbook
from Fano_Hartli import *
from openpyxl.styles import (
                        PatternFill, Border, Side, 
                        Alignment, Font, GradientFill
                        )


class Table(Workbook): 
    LANG = sorted('QWERTYUIOPASDFGHJKLMNBVCXZ')
    def __init__(self,stat:dict, sum_of_chars:int,*args, title:str='', show=False, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        print("Данные успешно получены")
        print(f"количесво символов в тексте: {sum_of_chars}")
        stat_ = list(stat.items())
        stat_.sort(key=lambda x: x[1], reverse=True)
        
        if show: 
            hartli = get_Hartli(stat_)
            print("Код Хартли успешно сформирован")
            # print(hartli.dict)
            haffman = get_Haffman(stat_)
            print("Код Хаффмана успешно сформирован")
            # print(haffman.dict)s

            hartli.save(filename="hartli.bmp")
            haffman.save(filename='haffman.bmp')
        
        hartli = hartli.dict
        haffman = haffman.dict
        
        self.ws1 = self.active
        self.ws1.title = "Часть1"
        self.ws2 = self.create_sheet("Часть2")
        self.ws3 = self.create_sheet("Часть3.1")
        self.ws3_2 = self.create_sheet("Часть3")
        self.ws2_1 = self.create_sheet("Хартли")
        self.ws3_3 = self.create_sheet("Хаффман")
        self.setws1(stat, ["ID", "Символ", "Код символа", "Число вхождений символа в текст", "Вероятность вхождения символа (рi)", "Ii"], sum_of_chars)
        self.setws2(stat, ["ID", "Символ", "Вероятность", "Код"], sum_of_chars)
        self.setws(self.ws2_1,hartli, ['шаг', 'комбинация', 'кол-во символов','символ'], sum_of_chars, 'Хартли')
        self.setws(self.ws3_3,haffman, ['шаг', 'комбинация', 'кол-во символов','символ'], sum_of_chars, 'Хаффман')
        self.setws3_1(stat, ["ID", "Символ", "Вероятность", "Код"], sum_of_chars)
        self.setws3_2(stat, ["ID", "Символы", "Вероятности"], sum_of_chars)
        

    @staticmethod
    def nfloat(a:float) -> str:
        return str(a).replace('.', ',')


    def setws1(self, stat:dict, header:list, sum_of_chars:int): 
        stat_ = list(stat.items())
        stat_.sort(key=lambda x: x[1])
        pv = 0 
        Isr = 0 
        
        self.ws1.append(header)
        for i in range(len(stat_)):
            char = stat_[i][0]
            count = stat_[i][1]
            self.ws1.append(list(map(str, (i+1, 
                                  char, 
                                  ord(char), 
                                  count, 
                                  self.nfloat(count/sum_of_chars), 
                                  self.nfloat(-math.log2(count/sum_of_chars))))))
            pv += count/sum_of_chars
            Isr += -math.log2(count/sum_of_chars) * count/sum_of_chars
            
        else: 
            self.ws1[f"C{i+3}"] = "Всего символов в тексте (K)"
            self.ws1[f"D{i+3}"] = sum_of_chars
            self.ws1[f"D{i+4}"] = "Полная вероятность(Р)"
            self.ws1[f"E{i+4}"] = self.nfloat(pv)
            self.ws1[f"E{i+5}"] = "Энтропия источника (Iср)"
            self.ws1[f"F{i+5}"] = self.nfloat(Isr)

    def setws(self, ws, codes:dict, header:list, sum_of_chars:int, type:str) -> None:
        ws.append(header)
        def calc_count(a:str, b:list) -> int: 
            count = 0
            for i in b: 
                if (len(a) < len(i)) and (a == i[:len(a)]): 
                    count += 1

            return count 
        word = 'У МЕНЯ ЕСТЬ ДЖИП ДЖИП'
        out = ''
        for i in word:
            out += codes[i]
        
        codes1 = dict(map(lambda x: x[::-1], codes.items()))
        res = ''
        i = 0
        print(f'{type}: {word} -> {out} -> {word}')
        count = 0
        while out: 
            i += 1
            count += 1
            key = out[:i]
            ls = [count, key]
            if key in codes1: 
                res += codes1[key]
                out = out[i:]
                ls += ['1', codes1[key]]
                i = 0
            else: 
                ls += [calc_count(key, list(codes1.keys())), '-']
            ws.append(ls)
            

        

    def get_norm_pos(self, a:int) -> str: 
        out = ''
        while a: 
            out += self.LANG[((a-1)%len(self.LANG))]
            a = (a-1)//len(self.LANG)
        return out[::-1]
        
    def setws2(self, stat:dict, header:list, sum_of_chars:int) -> None: 
        self.ws2.append(header)
        stat = list(stat.items())
        stat.sort(key=lambda x: x[1], reverse=True)
        
        codes = get_Hartli(stat)
        #codes.save("output")
        #codes.show()
        codes = codes.dict
        Isr = 0 
        stat = list(map(lambda x: (x[0], x[1]/sum_of_chars), stat))
        for i in range(len(stat)):
            char = stat[i][0]
            probab = stat[i][1]
            self.ws2.append([i+1, char, self.nfloat(probab), codes[char]])
            Isr += -math.log2(probab) * probab

        else: 
            self.ws2[f"A{i+3}"] = "Значение средней информации в битах"
            self.ws2[f"B{i+3}"] = self.nfloat(Isr)
        
        self.__set2(2, 50, 5)

    def __set2(self, a:int, b:int, column:int, deep:int=0): 
        for i in range(a, b): 
            code1 = self.ws2[f"D{i}"].value
            code2 = self.ws2[f"D{i+1}"].value
            
            if code1[deep] != code2[deep]: 
                double = Side(border_style="medium", color="000000")
                self.ws2[f"{self.get_norm_pos(column)}{i}"].border = Border(bottom=double)
                self.__set2(a, i, column+1, deep+1)
                self.__set2(i+1, b, column+1, deep+1)
        else: 
            return 

    def setws3_1(self, stat:dict, header:list, sum_of_chars:int) -> None: 
        self.ws3.append(header)
        stat = list(stat.items())
        stat.sort(key=lambda x: x[1], reverse=True)
        stat = list(map(lambda x: (x[0], x[1]/sum_of_chars), stat))
        codes = get_Haffman(stat)
        #codes.save('output2')
        #codes.show()
        codes = codes.dict

        Isr = 0 
        for i in range(len(stat)):
            char = stat[i][0]
            probab = stat[i][1]
            self.ws3.append([i+1, char, probab, codes[char]])
            Isr += -math.log2(probab) * probab
        else: 
            self.ws2[f"A{i+3}"] = "Значение средней информации в битах"
            self.ws2[f"B{i+3}"] = Isr

    def setws3_2(self, stat:dict, header:list, sum_of_chars:int) -> None:
        self.ws3_2.append(header)
        self.ws3_2.merge_cells('A1:A2')
        self.ws3_2.merge_cells('B1:B2')
        self.ws3_2.merge_cells('C1:C2')
        
        stat = list(stat.items())
        stat = list(map(lambda x: (x[0], x[1]/sum_of_chars), stat))
        stat.sort(key=lambda x: x[1], reverse=True)
        self.ws3_2.append([])
        for i in range(len(stat)):
            char = stat[i][0]
            probab = stat[i][1]
            self.ws3_2.append([i+1, char, self.nfloat(probab)])

        p = 4
        count = 0 
        while len(stat) > 1: 
            count += 1
            stat.sort(key=lambda x: x[1], reverse=True)
            stat = stat[:-2] + [(stat[-1][0] + stat[-2][0], stat[-1][1]+stat[-2][1])]
            p1 = self.get_norm_pos(p)
            p2 = self.get_norm_pos(p+1)
            self.ws3_2[f"{p2}{2}"] = f'Шаг\n{count}'
            for i in range(len(stat)):
                self.ws3_2[f'{p1}{i+3}'] = stat[i][0]
                self.ws3_2[f"{p2}{i+3}"] = stat[i][1]
            p += 2
        else: 
            self.ws3_2.merge_cells(f"{self.get_norm_pos(4)}1:{self.get_norm_pos(p)}1")
            self.ws3_2[f'{self.get_norm_pos(4)}1'] = 'Вспомагательные вычисления'
